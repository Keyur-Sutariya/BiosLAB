import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver import Keys
import openpyxl

current_row = 4
speciality_name_row = 4
speciality_details_row = 4

File = openpyxl.load_workbook("/home/tipl-pc-21/Desktop/Keyur/BIOS LAB/Excel/Data_Fetching/MD.xlsx")
active_file = File.active
last_row = active_file.max_row


driver = webdriver.Firefox()
driver.maximize_window()

actions = ActionChains(driver)

driver.get(" http://165.22.217.185/")
driver.find_element(By.XPATH, "//input[@placeholder='Please enter username'] ").send_keys("employee")
driver.find_element(By.XPATH, " //input[@placeholder='Please enter password']").send_keys("123123123")
driver.find_element(By.XPATH, " //button[@type='submit']").click()
driver.implicitly_wait(5)


driver.get("http://165.22.217.185/layout/master/doctor/Speciality")
driver.implicitly_wait(5)


while current_row <= last_row:

    speciality_stored = active_file.cell(row=speciality_name_row, column=2)
    speciality_name_stored = active_file.cell(row=speciality_details_row, column=3)

    speciality_stored_val = speciality_stored.value
    speciality_name_stored_val = speciality_name_stored.value
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//button[@class='btn btn-outline-primary mt-2 mx-2'] ").click()

    driver.find_element(By.XPATH, " //input[@name='SpecialityName']").send_keys("", speciality_stored_val)
    driver.find_element(By.XPATH, " //input[@name='Remarks']").send_keys("", speciality_name_stored_val)

    actions.send_keys(Keys.TAB).send_keys("MD").send_keys(Keys.ENTER).perform()
    driver.find_element(By.XPATH,"//button[@type='submit'] ").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/div/div[5]/div/div/div[1]/div[1]").click()
    driver.implicitly_wait(5)

    current_row = current_row+1
    speciality_name_row = speciality_name_row+1
    speciality_details_row = speciality_details_row+1


driver.quit()

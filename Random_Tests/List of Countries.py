import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from PIL import Image
from selenium.webdriver.common.action_chains import ActionChains


import time


driver = webdriver.Firefox()
driver.maximize_window()
driver.get("https://www.google.co.in")
action = ActionChains(driver)


driver.save_screenshot("Test2.png")
driver.get_screenshot_as_file("Test.png")

Country_File = openpyxl.load_workbook("/home/tipl-pc-21/Downloads/Countries.xlsx")
Active_File = Country_File.active
last_row = Active_File.max_row
print("Number of Rows: ", last_row)


current_row = 2
while current_row < last_row:
    stored_row = Active_File.cell(row=current_row, column=2)
    stored_row_value = stored_row.value
    print(stored_row_value)
    current_row = current_row+1


import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import Keys
from selenium.webdriver.common.action_chains import ActionChains

current_row = 2
email_row = 2
password_row = 2
Dashboard = "http://165.22.217.185/layout/dashboard"
File = openpyxl.load_workbook("/home/tipl-pc-21/Desktop/Usernames.xlsx")
Active_File = File.active
last_row = Active_File.max_row

print("Last row: ", last_row)
print("\n")

driver = webdriver.Firefox()
driver.get("http://165.22.217.185")
driver.maximize_window()
time.sleep(2)
action = ActionChains(driver)

while current_row < last_row:
    email_address = Active_File.cell(row=email_row, column=1)
    password = Active_File.cell(row=password_row, column=2)
    email_stored_value = email_address.value
    password_stored_value = password.value

    print("Email:    ", email_stored_value)
    print("Password: ", password_stored_value)

    driver.find_element("name", "LoginId").send_keys("", email_stored_value)
    driver.find_element("name", "Password").send_keys("", password_stored_value)
    driver.find_element(By.XPATH, "//div[button/@type='submit']").click()
    time.sleep(2)
    Current_page = driver.current_url

    if Current_page == Dashboard:
        driver.find_element(By.XPATH, "//button[@aria-label='close']//*[name()='svg']").click()
        print("Successfully Logged in")
        print("\n")
        email_row = email_row + 1
        password_row = password_row + 1
        current_row = current_row + 1
        time.sleep(2)
        driver.find_element(By.XPATH, "//span[@class='user-icon']").click()
        driver.find_element(By.XPATH, "//a[normalize-space()='Log Out']").click()
        driver.get("http://165.22.217.185")
        time.sleep(2)
    else:
        print("Wrong Username or Password")
        print("\n")
        driver.get_screenshot_as_file("Login_Failed.png")
        email_row = email_row + 1
        password_row = password_row + 1
        current_row = current_row + 1

        driver.get("http://165.22.217.185")
        time.sleep(2)

driver.quit()

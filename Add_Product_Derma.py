import time
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
import openpyxl

current_row = 44
Product_name_row = 44
Product_content_row = 44
Product_Short_Name_row = 44
i = 1
j = 7

driver = webdriver.Firefox()
File = openpyxl.load_workbook("/home/tipl-pc-21/Downloads/Product.xlsx")
Active_File = File.active
last_Row = Active_File.max_row
driver.get("http://165.22.217.185")
driver.find_element("name", "LoginId").send_keys("employee")
driver.find_element("name", "Password").send_keys("123123123")
driver.find_element(By.XPATH, "//div[button/@type='submit']").click()

time.sleep(2)

while current_row <= last_Row:
    i = i + 1
    Product_name_row_stored = Active_File.cell(row=Product_name_row, column=3)
    Product_content_row_stored = Active_File.cell(row=Product_content_row, column=4)
    Product_Short_Name_stored = Active_File.cell(row=Product_Short_Name_row, column=6)

    Product_content_row_stored_value = Product_content_row_stored.value
    Product_name_row_stored_value = Product_name_row_stored.value
    Product_Short_Name_stored_value = Product_Short_Name_stored.value

    driver.get("http://165.22.217.185/layout/AddProduct")
    time.sleep(1)
    action = ActionChains(driver)
    driver.find_element(By.XPATH, "//div[contains(@class,' css-1wy0on6')]//*[name()='svg']").click()
    time.sleep(1)
    action.key_down(Keys.DOWN).perform()
    action.key_down(Keys.ENTER).perform()
    action.key_up(Keys.DOWN).perform()
    action.key_up(Keys.ENTER).perform()
    time.sleep(1)

    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()
    action.send_keys(Keys.ENTER).perform()
    time.sleep(1)

    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()
    action.send_keys(Keys.ENTER).perform()
    time.sleep(1)
    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()
    action.send_keys(Keys.ENTER).perform()
    time.sleep(1)
    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()
    action.send_keys(Keys.ENTER).perform()
    time.sleep(1)
    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()
    action.send_keys(Keys.ENTER).perform()
    time.sleep(1)
    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()

    time.sleep(1)
    driver.find_element(By.XPATH, "//input[@name='Code']").send_keys(j, "964", i)
    time.sleep(1)
    action.send_keys(Keys.TAB).perform()
    driver.find_element(By.XPATH, " //input[@name='ShortName']").send_keys("", Product_Short_Name_stored_value)

    driver.find_element(By.XPATH, "//input[@name='Name'] ").send_keys("", Product_name_row_stored_value)
    time.sleep(1)
    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()
    action.send_keys(Keys.ENTER).perform()

    driver.find_element(By.XPATH, "//input[@name='CodeRef1']").send_keys("Reference Derma", i, j)

    driver.find_element(By.XPATH, "//input[@name='OnHold']").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "//input[@name='HsnCode']").send_keys(j, "541", i)
    time.sleep(1)
    action.send_keys(Keys.TAB).perform()
    action.send_keys(Keys.DOWN).perform()
    action.send_keys(Keys.ENTER).perform()
    time.sleep(1)
    driver.find_element(By.XPATH, "//input[@name='Cgst']").send_keys("9")
    time.sleep(1)
    driver.find_element(By.XPATH, "//input[@name='Sgst']").send_keys("")
    time.sleep(1)
    driver.find_element(By.XPATH, "//input[@name='Mrp']").send_keys("100")
    time.sleep(1)
    driver.find_element(By.XPATH, "//textarea[@name='Contents']").send_keys("", Product_content_row_stored_value)
    time.sleep(1)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()

    time.sleep(7)

    current_row = current_row + 1
    Product_name_row = Product_name_row + 1
    Product_content_row = Product_content_row + 1
    Product_Short_Name_row = Product_Short_Name_row + 1

    time.sleep(2)

# WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class,' css-1wy0on6')]//*[name()='svg']//svg/a/span[text()=' HO ']"))).click()
# select = Select(driver.find_element(By.XPATH, "//div[contains(@class,' css-1wy0on6')]//*[name()='svg']"))
# select.select_by_value(str("HO"))

driver.quit()
